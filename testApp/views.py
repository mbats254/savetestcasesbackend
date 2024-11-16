# products/views.py

from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import TestCase
from .serializers import TestCaseSerializer
import json
import os
from django.http import JsonResponse
from django.conf import settings
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

# List and Create View for TestCase
class TestCaseListView(APIView):
    """
    Retrieve a list of all test cases or create a new test case.
    """
    def get(self, request, format=None):
        test_cases = TestCase.objects.all()
        serializer = TestCaseSerializer(test_cases, many=True)
        return Response(serializer.data)

    def post(self, request, format=None):
        serializer = TestCaseSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# Detail, Update, and Delete View for a single TestCase
class TestCaseDetailView(APIView):
    """
    Retrieve, update, or delete a specific test case.
    """
    def get(self, request, pk, format=None):
        test_case = get_object_or_404(TestCase, pk=pk)
        serializer = TestCaseSerializer(test_case)
        return Response(serializer.data)

    def put(self, request, pk, format=None):
        test_case = get_object_or_404(TestCase, pk=pk)
        serializer = TestCaseSerializer(test_case, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk, format=None):
        test_case = get_object_or_404(TestCase, pk=pk)
        test_case.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
def load_all_test_cases(request):
    # Path to the JSON file
    json_file_path = os.path.join(settings.BASE_DIR, "data/data_with_ids.json")

    try:
        # Open and load the JSON data
        with open(json_file_path, "r") as file:
            test_cases = json.load(file)
        
        # Return data as JSON response
        return JsonResponse(test_cases, safe=False, status=200)

    except FileNotFoundError:
        return JsonResponse({"error": "File not found"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Error decoding JSON data"}, status=500)   


def download_testcases_excel_by_category(request):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Retrieve all unique categories from the TestCase model
    categories = TestCase.objects.values_list("category", flat=True).distinct()

    # Define the headers for the Excel sheets
    headers = ["ID", "Test Case", "Category", "Pre-condition", "Test Steps", "Test Data", "Expected Result", "Pass/Fail"]

    for category in categories:
        # Add a new worksheet for each category
        sheet = workbook.create_sheet(title=category if category else "Uncategorized")

        # Write the headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            sheet[f"{get_column_letter(col_num)}1"] = header

        # Retrieve test cases for the current category
        test_cases = TestCase.objects.filter(category=category)

        # Populate the worksheet with data from the TestCase model
        for row_num, test_case in enumerate(test_cases, start=2):
            sheet[f"A{row_num}"] = test_case.test_case_id
            sheet[f"B{row_num}"] = test_case.test_case
            sheet[f"C{row_num}"] = test_case.category
            sheet[f"D{row_num}"] = test_case.pre_condition
            sheet[f"E{row_num}"] = test_case.test_steps
            sheet[f"F{row_num}"] = test_case.test_data
            sheet[f"G{row_num}"] = test_case.expected_result
            
            # Fix for the pass_fail field, setting a default value if it's None or empty
            sheet[f"H{row_num}"] = test_case.pass_fail if test_case.pass_fail else "Not Set"

    # Remove the default sheet created by openpyxl if no data was added to it
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    # Set the response with the correct content type and filename for download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="test_cases_by_category.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response
    


def download_testcases_excel(request):
    # Create a new Excel workbook and a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TestCases"

    # Define the headers based on TestCase model fields
    headers = ["ID", "Test Case", "Category", "Pre-condition", "Test Steps", "Test Data", "Expected Result", "Pass/Fail"]
    for col_num, header in enumerate(headers, 1):
        sheet[f"{get_column_letter(col_num)}1"] = header

    # Retrieve all records from the TestCase model
    test_cases = TestCase.objects.all()

    # Populate the Excel sheet with data from the TestCase model
    for row_num, test_case in enumerate(test_cases, start=2):
        sheet[f"A{row_num}"] = test_case.test_case_id
        sheet[f"B{row_num}"] = test_case.test_case
        sheet[f"C{row_num}"] = test_case.category
        sheet[f"D{row_num}"] = test_case.pre_condition
        sheet[f"E{row_num}"] = test_case.test_steps
        sheet[f"F{row_num}"] = test_case.test_data
        sheet[f"G{row_num}"] = test_case.expected_result
        sheet[f"H{row_num}"] = test_case.pass_fail

    # Set the response with the correct content type and filename for download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="test_cases.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response   


def download_testcases_json(request):
    # Retrieve all records from the TestCase model
    test_cases = TestCase.objects.all()

    # Prepare data in JSON format with specified fields
    data = []
    for test_case in test_cases:
        data.append({
            "id": test_case.test_case_id,  # Assuming the model has this field
            "__EMPTY": "",  # Placeholder field, as per your structure
            "Test Case": test_case.test_case,
            "Pre-condition": test_case.pre_condition,
            "Test Steps": test_case.test_steps,
            "Test Data": test_case.test_data,
            "Expected Result": test_case.expected_result,
            "Actual Result": "",  # Placeholder if not provided by the model
            "Device": "",         # Placeholder if not provided by the model
            "Pass/Fail": test_case.pass_fail,
            "Bug(Status)": ""     # Placeholder if not provided by the model
        })

    # Convert the data to JSON format
    json_data = json.dumps(data, indent=4)

    # Set the response with the correct content type and filename for download
    response = HttpResponse(
        json_data,
        content_type="application/json"
    )
    response["Content-Disposition"] = 'attachment; filename="data_with_ids.json"'

    return response

    
